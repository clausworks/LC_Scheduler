from utils import get_all_shifts, get_all_subjects_dict
from tutor import Shift
from openpyxl.styles import Font, Alignment


def write_shift_names(ws, tutors):
    # Generate headers.
    all_shifts = get_all_shifts(tutors)
    for i,shift in enumerate(all_shifts):
        ws.cell(row=i+2, column=1).value = shift
        ws.cell(row=i+2, column=1).font = Font(bold=True)
        ws.cell(row=i+2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    days = Shift.DAYS
    for j,day in enumerate(days):
        ws.cell(row=1, column=j+2).value = day
        ws.cell(row=1, column=j+2).font = Font(bold=True)
        ws.cell(row=1, column=j+2).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].width = 10


    # Place names.
    for i,cur_shift in enumerate(all_shifts):
        for j,cur_day in enumerate(days):
            # Format cells.
            ws.cell(row=i+2, column=j+2).alignment = Alignment(wrapText=True,
                                                     horizontal='center',
                                                     vertical='center')
            # FIXME: adjust width too
            # Place text.
            for tutor in tutors:
                for tutor_shift in tutor.shifts:
                    if cur_shift == tutor_shift.name and cur_day == tutor_shift.day:
                        if (tutor.fname + ' ' + tutor.lname[:1]) in str(ws.cell(row=i+2, column=j+2).value):
                            print('Duplicate:', tutor.fname + ' ' + tutor.lname[:1])
                            for x in tutor.shifts:
                                print(x)
                        if ws.cell(row=i+2, column=j+2).value:
                            ws.cell(row=i+2, column=j+2).value += (tutor.fname + ' '
                                    + tutor.lname[:1] + '.\n')
                        else:
                            ws.cell(row=i+2, column=j+2).value = (tutor.fname + ' '
                                    + tutor.lname[:1] + '.\n')


def write_shift_subjects(ws, tutors):
    '''
    Generates a shifts vs. names table in the given worksheet
    using data from the given tutors.
    '''

    all_shifts = get_all_shifts(tutors)
    all_subjects = get_all_subjects_dict(tutors)

    # Write header column (cell coordinates: row=i,col=j)

    # Write header at top.
    j = 3 # Column index for current cell.
    for cur_day in Shift.DAYS:
        ws.cell(row=1, column=j).value = cur_day  # DAY
        for cur_shift in all_shifts:
            ws.cell(row=2, column=j).value = cur_shift  # Shift N
            j += 1

    # Write header on left side.
    i = 3  # Row index for current subject
    for subj_area,nums in all_subjects.items():
        ws.cell(row=i, column=1).value = subj_area
        for m,cur_num in enumerate(nums):
            ws.cell(row=i, column=2).value = cur_num
            i += 1

    # Write data.
    last_valid_subj_area = None
    for i,row in enumerate(ws.iter_rows()):
        if row[0].value:
            cur_subj_area = row[0].value
            last_valid_subj_area = cur_subj_area
        else:
            cur_subj_area = last_valid_subj_area

        for j in range(len(row)):
            for tutor in tutors:
                if tutor.subjects:
                    for tutor_subj in tutor.subjects:
                        for tutor_shift in tutor.shifts:
                            # Round down to nearest multiple of the number of
                            # shifts to get the index of the current day.
                            if (i>1 and j>1):
                                day_col = ((j+2) // len(all_shifts)) * len(all_shifts)

                                if (tutor_subj.name == last_valid_subj_area
                                    and tutor_subj.number == row[1].value
                                    and tutor_shift.day == ws.cell(row=1, column=day_col).value
                                    and tutor_shift.name == ws.cell(row=2, column=j+2).value):
                                    row[j].value = 'X'


def write_subject_names(ws, tutors):
    all_subjects = get_all_subjects_dict(tutors)

    i=1
    for subj_area,nums in all_subjects.items():
        ws.cell(row=i, column=1).value = subj_area
        for cur_num in nums:
            ws.cell(row=i, column=2).value = cur_num
            # TODO: Write this to a list and sort the list before writing to cell.
            for tutor in tutors:
                if tutor.subjects:
                    for tutor_subj in tutor.subjects:
                        if (tutor_subj.name == subj_area
                            and tutor_subj.number == cur_num):
                            if ws.cell(row=i, column=3).value == None:
                                ws.cell(row=i, column=3).value = str(tutor)
                            else:
                                ws.cell(row=i, column=3).value += ', ' + str(tutor)
            i += 1
