#!/usr/bin/python3
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from tutor import Tutor
from collectors import read_name_shifts, read_name_subjects
from generators import write_shift_names, write_shift_subjects, write_subject_names
from utils import get_writing_tutors, get_nonwriting_tutors, get_all_subjects_dict
import os
import sys

def main():
    wb_filename = 'schedule.xlsx'
    try:
        # determine if application is a script file or frozen exe
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)
        wb_path = os.path.join(application_path, wb_filename)
        wb = load_workbook(filename=wb_path, read_only=True)
    except FileNotFoundError:
        print('-'*80)
        print('The Excel workbook schedule.xlsx could not be loaded.')
        print('Please make sure to name the file you want to use "schedule.xlsx"')
        print('-'*80)
        return

    err_file = open('ERRORS.txt', 'w')  # Clear/create error file.

    # Read data into dictionaries.
    try:
        name_shifts = read_name_shifts(wb['Tutor Schedule'])
        name_subjects = read_name_subjects(wb['Subjects'])
    except KeyError:
        print('-'*80)
        print('The workbook schedule.xlsx is missing the following worksheets:')
        if 'Tutor Schedule' not in wb:
            print('\tTutor Schedule')
        if 'Subjects' not in wb:
            print('\tSubjects')
        print('-'*80)
        return

    err_output = []
    for name in name_subjects:
        if name not in name_shifts:
            err_output.append(name)
    if err_output:
        err_file = open('ERRORS.txt', 'a')
        print('The following tutors have subjects associated with their names', file=err_file)
        print('but they do not have any shifts associated with their names:', file=err_file)
        for i,tutor_name in enumerate(err_output):
            print('\t', i+1, ': ',
                tutor_name[0],', ',tutor_name[1],
                sep='', file=err_file)
        print('',file=err_file)
        err_file.close()

    # Generate tutor objects from data dictionaries.
    tutors = []
    for name in name_shifts:
        tutor = Tutor(lname=name[0],
                      fname=name[1],
                      shifts=name_shifts[name])
        tutors.append(tutor)

    # Note: Writing tutors don't have subjects.
    writing_tutors = get_writing_tutors(tutors)

    # Update GT/Math tutors with their subjects.
    err_output = []
    nonwriting_tutors = get_nonwriting_tutors(tutors)
    for tutor in nonwriting_tutors:
        cur_name = (tutor.lname,tutor.fname)
        if cur_name in name_subjects:
            tutor.subjects = name_subjects[cur_name]
        else:
            # No subjects listed for this tutor.
            err_output.append(tutor.lname + ', ' + tutor.fname)
    if err_output:
        err_file = open('ERRORS.txt', 'a')
        print('The following GT/Math tutors have shifts associated with their names,', file=err_file)
        print('but they do not have any subjects associated with their names:', file=err_file)
        for i,tutor_name in enumerate(err_output):
            print('\t', i+1, ': ',
                tutor_name,
                sep='', file=err_file)
        print('',file=err_file)
        err_file.close()

    # Log tutors to file.
    tutor_file = open('TUTORS.txt', 'w')  # Create/overwrite file.
    print('WRITING TUTORS', file=tutor_file)
    for tutor in writing_tutors:
        print('\t', sep='', end='', file=tutor_file)
        print(tutor.fname, tutor.lname, file=tutor_file)
    print('\nGT/MATH TUTORS', file=tutor_file)
    for tutor in nonwriting_tutors:
        print('\t', sep='', end='', file=tutor_file)
        print(tutor.fname, tutor.lname, file=tutor_file)
    print('\nUNCLASSIFIED TUTORS', file=tutor_file)
    print('These tutors do not have valid shifts associated with their names.', file=tutor_file)
    print('This could be because they simply do not have any drop-in shifts,', file=tutor_file)
    print('but it also could be because their drop-in shifts have not been', file=tutor_file)
    print('updated to the new shift format.', file=tutor_file)
    for tutor in tutors:
        if tutor not in writing_tutors and tutor not in nonwriting_tutors:
            print('\t', sep='', end='', file=tutor_file)
            print(tutor.fname, tutor.lname, file=tutor_file)
    tutor_file.close()

    # Write to workbook.
    new_wb = Workbook()

    ws_gtma_name = 'GT & Math (by Tutor)'
    new_wb.active.title = ws_gtma_name  # Rename default new sheet.
    write_shift_names(new_wb[ws_gtma_name], nonwriting_tutors)

    ws_gtma_subj_sched_name = 'GT & Math (by Subject)'
    new_wb.create_sheet(ws_gtma_subj_sched_name)
    write_shift_subjects(new_wb[ws_gtma_subj_sched_name], nonwriting_tutors)

    ws_subjects_name = 'Subjects & Tutors'
    new_wb.create_sheet(ws_subjects_name)
    write_subject_names(new_wb[ws_subjects_name], nonwriting_tutors)

    new_wb.save('Generated Schedules.xlsx')
    print('-'*80)
    print('The new schedules have been successfully generated.')
    print('You can locate them as worksheets in the Generated Schedules workbook.')
    print('-'*80)

if __name__ == '__main__':
    main()
    input('\nPress ENTER to close this window.')
